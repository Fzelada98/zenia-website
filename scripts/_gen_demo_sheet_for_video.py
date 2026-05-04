"""
Genera el xlsx demo COMPLETO para grabar el Loom.
Marca: OLIVAR DEL SUR (importadora ficticia de aceite de oliva)
- 6 pestanas: Portada, Dashboard, Transacciones, Cashflow, Deudas, KPIs
- Logo embebido en Portada y Dashboard
- 30 transacciones con pagos proveedores + cobros clientes + pagos deuda
- Cashflow diario + grafico LineChart (sin gridlines, fechas legibles)
- 3 buckets calculados con formulas
- Hero stat + alertas
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import LineChart, Reference
import os

# Brand palette OLIVAR DEL SUR
DARK_BG = "1F4D2E"
GOLD = "C9A961"
GOLD_DARK = "8B6914"
GOLD_LIGHT = "F5ECD7"
LINK_COLOR = "059669"
WHITE = "FFFFFF"
TEXT_PRIMARY = "1A1A1A"
TEXT_MUTED = "94A3B8"
GREEN_BG = "DCFCE7"
GREEN_TEXT = "166534"
RED_BG = "FEE2E2"
RED_TEXT = "991B1B"
BLUE_BG = "DBEAFE"
BLUE_TEXT = "2563EB"
AMBER_BG = "FEF3C7"
AMBER_TEXT = "92400E"
ORANGE_BG = "FED7AA"
ORANGE_TEXT = "9A3412"

wb = Workbook()
thin = Side(border_style="thin", color="E2E8F0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def fill(color):
    return PatternFill("solid", fgColor=color)


LOGO_PATH = os.path.join("posts", "olivar-del-sur-logo.png")


# =========================
# 1. PORTADA (cover only)
# =========================
ws = wb.active
ws.title = "Portada"
ws.sheet_view.showGridLines = False

for col, w in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I'],
                  [3, 22, 18, 22, 18, 22, 18, 22, 22]):
    ws.column_dimensions[col].width = w

# Logo grande centrado
if os.path.exists(LOGO_PATH):
    img = XLImage(LOGO_PATH)
    img.width = 480
    img.height = 144
    img.anchor = "C3"
    ws.add_image(img)
for r in range(1, 9):
    ws.row_dimensions[r].height = 22

# Tagline
ws['B11'] = "REPOSITORIO ZELLE  ·  CONTROL FINANCIERO 360"
ws['B11'].font = Font(size=18, bold=True, color=GOLD)
ws['B11'].fill = fill(DARK_BG)
ws.merge_cells('B11:I11')
ws['B11'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[11].height = 42

ws['B12'] = "Mes en curso: Mayo 2026  ·  Demo creado por ZENIA  ·  zeniapartners.com"
ws['B12'].font = Font(size=11, italic=True, color="64748B")
ws.merge_cells('B12:I12')
ws['B12'].alignment = Alignment(horizontal='center')
ws.row_dimensions[12].height = 22

ws['B13'] = "Datos ficticios. Cliente, montos y movimientos inventados."
ws['B13'].font = Font(size=9, italic=True, color="94A3B8")
ws.merge_cells('B13:I13')
ws['B13'].alignment = Alignment(horizontal='center')

# Indice de pestanas
ws['B16'] = "INDICE DE PESTAÑAS"
ws['B16'].font = Font(size=13, bold=True, color=GOLD)
ws['B16'].fill = fill(DARK_BG)
ws.merge_cells('B16:I16')
ws['B16'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws.row_dimensions[16].height = 28

index_items = [
    ("Dashboard",     "Buckets de cash, KPIs, alertas, top pagadores, cashflow visual"),
    ("Transacciones", "Repositorio de movimientos. La asistente escribe aqui cada captura"),
    ("Cashflow",      "Detalle diario con grafico de saldo acumulado del mes"),
    ("Deudas",        "Prestamos por pagar, intereses mensuales, credito rotativo proveedor"),
    ("KPIs",          "Referencia tecnica de las formulas usadas"),
]
for i, (name, desc) in enumerate(index_items, 18):
    ws.cell(row=i, column=2, value=name).font = Font(size=12, bold=True, color=DARK_BG)
    ws.cell(row=i, column=2).fill = fill(GOLD_LIGHT)
    ws.cell(row=i, column=2).alignment = Alignment(horizontal='left', indent=1, vertical='center')
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
    ws.cell(row=i, column=4, value=desc).font = Font(size=10, color=TEXT_PRIMARY)
    ws.cell(row=i, column=4).alignment = Alignment(horizontal='left', indent=1, vertical='center')
    ws.merge_cells(start_row=i, start_column=4, end_row=i, end_column=9)
    ws.row_dimensions[i].height = 26

# Footer
ws['B26'] = "Asistente IA en WhatsApp  ·  Powered by Zenia Partners"
ws['B26'].font = Font(size=10, italic=True, color="64748B")
ws.merge_cells('B26:I26')
ws['B26'].alignment = Alignment(horizontal='center')


# =========================
# 2. DASHBOARD (live data)
# =========================
ws_d = wb.create_sheet("Dashboard")
ws_d.sheet_view.showGridLines = False

for col, w in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I'],
                  [3, 22, 18, 22, 18, 22, 18, 22, 22]):
    ws_d.column_dimensions[col].width = w

# Logo top-left + title row
if os.path.exists(LOGO_PATH):
    img = XLImage(LOGO_PATH)
    img.width = 280
    img.height = 84
    img.anchor = "B1"
    ws_d.add_image(img)
ws_d.row_dimensions[1].height = 32
ws_d.row_dimensions[2].height = 32
ws_d.row_dimensions[3].height = 32

ws_d['F2'] = "DASHBOARD OPERATIVO"
ws_d['F2'].font = Font(size=18, bold=True, color=DARK_BG)
ws_d.merge_cells('F2:I2')
ws_d['F2'].alignment = Alignment(horizontal='right', vertical='center')

ws_d['F3'] = "Mes en curso · Mayo 2026"
ws_d['F3'].font = Font(size=11, italic=True, color="64748B")
ws_d.merge_cells('F3:I3')
ws_d['F3'].alignment = Alignment(horizontal='right', vertical='center')

# Saldo inicial editable
ws_d['B5'] = "SALDO INICIAL BANCO (editable):"
ws_d['B5'].font = Font(size=10, italic=True, color="64748B")
ws_d.merge_cells('B5:D5')
ws_d['B5'].alignment = Alignment(horizontal='right')
ws_d['E5'] = 35000
ws_d['E5'].font = Font(size=11, bold=True, color=BLUE_TEXT)
ws_d['E5'].fill = fill(BLUE_BG)
ws_d['E5'].number_format = '"$"#,##0.00'

# Hero stat
ws_d['B7'] = "EFECTIVO TRAZADO TOTAL"
ws_d['B7'].font = Font(size=11, bold=True, color=GOLD)
ws_d['B7'].fill = fill(DARK_BG)
ws_d.merge_cells('B7:I7')
ws_d['B7'].alignment = Alignment(horizontal='center', vertical='center')
ws_d.row_dimensions[7].height = 22

ws_d['B8'] = "=C12+E12+G12"
ws_d['B8'].font = Font(size=36, bold=True, color=DARK_BG)
ws_d['B8'].fill = fill(GOLD_LIGHT)
ws_d['B8'].number_format = '"$"#,##0.00'
ws_d.merge_cells('B8:I8')
ws_d['B8'].alignment = Alignment(horizontal='center', vertical='center')
ws_d.row_dimensions[8].height = 56

ws_d['B9'] = "Suma trazada: liquido en banco + capital en gandolas + cuentas por cobrar"
ws_d['B9'].font = Font(size=10, italic=True, color="64748B")
ws_d.merge_cells('B9:I9')
ws_d['B9'].alignment = Alignment(horizontal='center')

# Trazabilidad header
ws_d['B11'] = "TRAZABILIDAD DEL EFECTIVO"
ws_d['B11'].font = Font(size=12, bold=True, color=GOLD)
ws_d['B11'].fill = fill(DARK_BG)
ws_d.merge_cells('B11:I11')
ws_d['B11'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws_d.row_dimensions[11].height = 0  # collapse: header is on same row as cards

# Re-do: place trazabilidad header at row 10 separator and cards at 11-13
ws_d['B11'] = None
ws_d.unmerge_cells('B11:I11')
ws_d.row_dimensions[11].height = 22

cards = [
    ('B', 'C', 'EFECTIVO EN BANCO',
     '=E5+SUMIFS(Transacciones!D:D,Transacciones!E:E,"in")-SUMIFS(Transacciones!D:D,Transacciones!E:E,"out")-(SUMIFS(Transacciones!D:D,Transacciones!F:F,"entrega_gandola")-SUMIFS(Transacciones!D:D,Transacciones!F:F,"retorno_gandola"))-(SUMIFS(Transacciones!D:D,Transacciones!F:F,"venta_credito")-SUMIFS(Transacciones!D:D,Transacciones!F:F,"cobro_credito"))',
     'Saldo liquido', BLUE_BG),
    ('D', 'E', 'DINERO EN GANDOLAS',
     '=SUMIFS(Transacciones!D:D,Transacciones!F:F,"entrega_gandola")-SUMIFS(Transacciones!D:D,Transacciones!F:F,"retorno_gandola")',
     'Capital en transito', AMBER_BG),
    ('F', 'G', 'CUENTAS POR COBRAR',
     '=SUMIFS(Transacciones!D:D,Transacciones!F:F,"venta_credito")-SUMIFS(Transacciones!D:D,Transacciones!F:F,"cobro_credito")',
     'Pendiente de cobrar', ORANGE_BG),
    ('H', 'I', 'TOTAL TRAZADO', '=C12+E12+G12', 'Suma de los 3', GREEN_BG),
]
for c_left, c_right, label, formula, sub, bg in cards:
    rng = f"{c_left}11:{c_right}11"
    ws_d.merge_cells(rng)
    ws_d[f"{c_left}11"] = label
    ws_d[f"{c_left}11"].font = Font(size=10, bold=True, color=DARK_BG)
    ws_d[f"{c_left}11"].fill = fill(bg)
    ws_d[f"{c_left}11"].alignment = Alignment(horizontal='center', vertical='center')

    rng2 = f"{c_left}12:{c_right}12"
    ws_d.merge_cells(rng2)
    ws_d[f"{c_left}12"] = formula
    ws_d[f"{c_left}12"].font = Font(size=20, bold=True, color=DARK_BG)
    ws_d[f"{c_left}12"].fill = fill(bg)
    ws_d[f"{c_left}12"].alignment = Alignment(horizontal='center', vertical='center')
    ws_d[f"{c_left}12"].number_format = '"$"#,##0.00'

    rng3 = f"{c_left}13:{c_right}13"
    ws_d.merge_cells(rng3)
    ws_d[f"{c_left}13"] = sub
    ws_d[f"{c_left}13"].font = Font(size=9, italic=True, color="64748B")
    ws_d[f"{c_left}13"].fill = fill(bg)
    ws_d[f"{c_left}13"].alignment = Alignment(horizontal='center')

ws_d.row_dimensions[12].height = 60
ws_d.row_dimensions[13].height = 18

# KPIs
ws_d['B15'] = "KPIs DEL PERIODO"
ws_d['B15'].font = Font(size=12, bold=True, color=GOLD)
ws_d['B15'].fill = fill(DARK_BG)
ws_d.merge_cells('B15:I15')
ws_d['B15'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws_d.row_dimensions[15].height = 26

kpi_cards = [
    ('B', 'C', 'INGRESOS COBRADOS', '=SUMIFS(Transacciones!D:D,Transacciones!E:E,"in")', 'Total in', GREEN_BG),
    ('D', 'E', 'GASTOS PAGADOS', '=SUMIFS(Transacciones!D:D,Transacciones!E:E,"out")', 'Total out', RED_BG),
    ('F', 'G', 'FLUJO NETO', '=C18-E18', 'Cashflow del periodo', GOLD_LIGHT),
    ('H', 'I', 'TICKET PROMEDIO', '=IFERROR(C18/COUNTIFS(Transacciones!E:E,"in"),0)', 'Por cobro', BLUE_BG),
]
for c_left, c_right, label, formula, sub, bg in kpi_cards:
    rng = f"{c_left}17:{c_right}17"
    ws_d.merge_cells(rng)
    ws_d[f"{c_left}17"] = label
    ws_d[f"{c_left}17"].font = Font(size=10, bold=True, color=DARK_BG)
    ws_d[f"{c_left}17"].fill = fill(bg)
    ws_d[f"{c_left}17"].alignment = Alignment(horizontal='center', vertical='center')

    rng2 = f"{c_left}18:{c_right}18"
    ws_d.merge_cells(rng2)
    ws_d[f"{c_left}18"] = formula
    ws_d[f"{c_left}18"].font = Font(size=20, bold=True, color=LINK_COLOR)
    ws_d[f"{c_left}18"].fill = fill(bg)
    ws_d[f"{c_left}18"].alignment = Alignment(horizontal='center', vertical='center')
    ws_d[f"{c_left}18"].number_format = '"$"#,##0.00'

    rng3 = f"{c_left}19:{c_right}19"
    ws_d.merge_cells(rng3)
    ws_d[f"{c_left}19"] = sub
    ws_d[f"{c_left}19"].font = Font(size=9, italic=True, color="64748B")
    ws_d[f"{c_left}19"].fill = fill(bg)
    ws_d[f"{c_left}19"].alignment = Alignment(horizontal='center')

ws_d.row_dimensions[17].height = 22
ws_d.row_dimensions[18].height = 60
ws_d.row_dimensions[19].height = 18

# Alertas
ws_d['B21'] = "ALERTAS Y RESUMEN OPERATIVO"
ws_d['B21'].font = Font(size=12, bold=True, color=GOLD)
ws_d['B21'].fill = fill(DARK_BG)
ws_d.merge_cells('B21:I21')
ws_d['B21'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws_d.row_dimensions[21].height = 26

alerts = [
    ('B', 'D', '!  GANDOLAS ACTIVAS', '=COUNTIFS(Transacciones!F:F,"entrega_gandola")-COUNTIFS(Transacciones!F:F,"retorno_gandola")', 'Capital en ruta sin cerrar', AMBER_BG, AMBER_TEXT, '0" gandolas"'),
    ('E', 'F', '!  PROXIMO PAGO INTERES', 1600, 'Luis Mendoza · vence 28-MAY', RED_BG, RED_TEXT, '"$"#,##0.00'),
    ('G', 'I', '+  COBROS DEL PERIODO', '=COUNTIFS(Transacciones!E:E,"in")', 'Pagos recibidos', GREEN_BG, GREEN_TEXT, '0" cobros"'),
]
for c_left, c_right, label, formula, sub, bg, text_color, fmt in alerts:
    rng = f"{c_left}23:{c_right}23"
    ws_d.merge_cells(rng)
    ws_d[f"{c_left}23"] = label
    ws_d[f"{c_left}23"].font = Font(size=10, bold=True, color=text_color)
    ws_d[f"{c_left}23"].fill = fill(bg)
    ws_d[f"{c_left}23"].alignment = Alignment(horizontal='center', vertical='center')

    rng2 = f"{c_left}24:{c_right}24"
    ws_d.merge_cells(rng2)
    ws_d[f"{c_left}24"] = formula
    ws_d[f"{c_left}24"].font = Font(size=22, bold=True, color=text_color)
    ws_d[f"{c_left}24"].fill = fill(bg)
    ws_d[f"{c_left}24"].alignment = Alignment(horizontal='center', vertical='center')
    ws_d[f"{c_left}24"].number_format = fmt

    rng3 = f"{c_left}25:{c_right}25"
    ws_d.merge_cells(rng3)
    ws_d[f"{c_left}25"] = sub
    ws_d[f"{c_left}25"].font = Font(size=9, italic=True, color="64748B")
    ws_d[f"{c_left}25"].fill = fill(bg)
    ws_d[f"{c_left}25"].alignment = Alignment(horizontal='center')

ws_d.row_dimensions[23].height = 22
ws_d.row_dimensions[24].height = 50
ws_d.row_dimensions[25].height = 18

# Top pagadores
ws_d['B27'] = "TOP PAGADORES DEL PERIODO"
ws_d['B27'].font = Font(size=12, bold=True, color=GOLD)
ws_d['B27'].fill = fill(DARK_BG)
ws_d.merge_cells('B27:I27')
ws_d['B27'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws_d.row_dimensions[27].height = 26

ws_d['B28'] = "Cliente"
ws_d['D28'] = "Total Cobrado"
ws_d['F28'] = "# Pagos"
ws_d['H28'] = "Ticket Promedio"
for cell in ['B28', 'D28', 'F28', 'H28']:
    ws_d[cell].font = Font(size=10, bold=True, color=DARK_BG)
    ws_d[cell].fill = fill(GOLD_LIGHT)
    ws_d[cell].alignment = Alignment(horizontal='center')

top_clients = [
    ("Empresa XYZ",       1700.00, 2, 850.00),
    ("Lucia Ramirez",     1300.00, 2, 650.00),
    ("Carlos Perez",      1200.00, 2, 600.00),
    ("Pedro Lopez",        800.00, 2, 400.00),
    ("Carmona (cliente)",  650.00, 1, 650.00),
    ("Maria Gonzalez",     500.00, 1, 500.00),
]
for i, (name, total, count, avg) in enumerate(top_clients, 29):
    ws_d.cell(row=i, column=2, value=name).font = Font(size=10, bold=True, color=TEXT_PRIMARY)
    ws_d.cell(row=i, column=4, value=total).font = Font(size=11, bold=True, color=GREEN_TEXT)
    ws_d.cell(row=i, column=4).number_format = '"$"#,##0.00'
    ws_d.cell(row=i, column=4).alignment = Alignment(horizontal='right')
    ws_d.cell(row=i, column=6, value=count).font = Font(size=10, color=TEXT_PRIMARY)
    ws_d.cell(row=i, column=6).alignment = Alignment(horizontal='center')
    ws_d.cell(row=i, column=8, value=avg).font = Font(size=10, color=TEXT_PRIMARY)
    ws_d.cell(row=i, column=8).number_format = '"$"#,##0.00'
    ws_d.cell(row=i, column=8).alignment = Alignment(horizontal='right')


# =========================
# 3. TRANSACCIONES (30 rows: cobros + pagos proveedores + pagos deuda)
# =========================
ws2 = wb.create_sheet("Transacciones")
ws2.sheet_view.showGridLines = False

headers = ["ID", "Fecha", "Nombre", "Monto", "Direccion", "Categoria", "Conf#", "Estado", "Notas"]
for i, h in enumerate(headers, 1):
    c = ws2.cell(row=1, column=i, value=h)
    c.font = Font(bold=True, color=GOLD, size=11)
    c.fill = fill(DARK_BG)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border
ws2.row_dimensions[1].height = 32

demo_txs = [
    # MAGIC ROW: matchea captura zelle-fake-01-maria.png
    ("D-001", "2026-05-04", "Maria Gonzalez",        500.00, "in",  "cobro_cliente",       "ABC123XYZ", "completed", "Pago anticipo orden #4421"),
    ("D-002", "2026-05-03", "Carlos Perez",          700.00, "in",  "cobro_cliente",       "XYZ789DEF", "completed", "Pago final pedido botellas"),
    ("D-003", "2026-05-03", "Empresa XYZ",           850.00, "in",  "cobro_cliente",       "QWE456RTY", "completed", "Pago factura mayo"),
    ("D-004", "2026-05-03", "Olivar Andaluz S.L.",  3500.00, "out", "pago_proveedor",      "WT26050301", "completed", "Pago proveedor aceite Espana"),
    ("D-005", "2026-05-02", "Carlos Perez",          500.00, "in",  "cobro_cliente",       "MNB789POI", "completed", "Anticipo pedido grande"),
    ("D-006", "2026-05-02", "Lucia Ramirez",         650.00, "in",  "cobro_cliente",       "ZXC123ASD", "completed", "Pago semanal restaurante"),
    ("D-007", "2026-05-02", "Empacadora Industrial", 1200.00, "out", "pago_proveedor",      "WT26050202", "completed", "Pago envases / etiquetas"),
    ("D-008", "2026-05-01", "Gandola Beleva",       5000.00, "out", "entrega_gandola",     "",          "completed", "Capital viaje Caracas"),
    ("D-009", "2026-05-01", "Pedro Lopez",           400.00, "in",  "cobro_cliente",       "WER567TYU", "completed", "Pago tienda esquina"),
    ("D-010", "2026-04-30", "Juan Rodriguez",        800.00, "out", "venta_credito",       "",          "completed", "Mercancia paga el 15 de mayo"),
    ("D-011", "2026-04-30", "Planilla Equipo",      2500.00, "out", "planilla",            "",          "completed", "Quincena fin abril"),
    ("D-012", "2026-04-29", "Aduana DAE",           1350.00, "out", "impuesto_gandola",    "",          "completed", "Permiso ingreso Venezuela"),
    ("D-013", "2026-04-29", "Pedro Ramos",          5000.00, "out", "pago_deuda_capital",  "WT26042901", "completed", "Abono a capital prestamo (capital baja a $45K)"),
    ("D-014", "2026-04-28", "Luis Mendoza",         1600.00, "out", "pago_deuda_interes",  "",          "completed", "Interes mensual abril (4%)"),
    ("D-015", "2026-04-28", "Pedro Ramos",          2500.00, "out", "pago_deuda_interes",  "",          "completed", "Interes mensual abril (5%)"),
    ("D-016", "2026-04-27", "Logistica Andes",      2800.00, "out", "gasto_logistica",     "WT26042701", "completed", "Flete contenedor #C-2204"),
    ("D-017", "2026-04-26", "Gandola Beleva",       4200.00, "in",  "retorno_gandola",     "",          "completed", "Cierre viaje Caracas - efectivo"),
    ("D-018", "2026-04-25", "Viaticos chofer",       320.00, "out", "viaticos",            "",          "completed", "Comida + peajes Caracas"),
    ("D-019", "2026-04-25", "Lucia Ramirez",         650.00, "in",  "cobro_cliente",       "POI234UYT", "completed", "Pago semanal restaurante"),
    ("D-020", "2026-04-24", "Empresa XYZ",           850.00, "in",  "cobro_cliente",       "LKJ567HGF", "completed", "Pago factura abril"),
    ("D-021", "2026-04-24", "Almacen Doral",         780.00, "out", "alquiler",            "",          "completed", "Renta mayo galpon"),
    ("D-022", "2026-04-23", "Pedro Lopez",           400.00, "in",  "cobro_cliente",       "MNB456VCX", "completed", "Pago semanal"),
    ("D-023", "2026-04-23", "Botellas Premium S.A.", 1850.00, "out", "pago_proveedor",      "WT26042302", "completed", "Pago vidrio 1L (5,000 botellas)"),
    ("D-024", "2026-04-22", "Restaurante Lima",      450.00, "in",  "cobro_cliente",       "QAZ123WSX", "completed", "Pago primera caja"),
    ("D-025", "2026-04-22", "Combustible flota",     240.00, "out", "combustible",         "",          "completed", "Gasolina 2 furgonetas"),
    ("D-026", "2026-04-21", "Gandola Beleva",       4500.00, "out", "entrega_gandola",     "",          "completed", "Capital nuevo viaje Maracaibo"),
    ("D-027", "2026-04-20", "Asesoria Contable",     480.00, "out", "honorarios",          "",          "completed", "Honorarios mensuales"),
    ("D-028", "2026-04-20", "Pedro Ramos",          2500.00, "out", "pago_deuda_interes",  "",          "completed", "Interes mensual marzo (5%) atrasado"),
    ("D-029", "2026-04-19", "Carmona (cliente)",     650.00, "in",  "cobro_credito",       "VBN234ZAQ", "completed", "Cancela mercancia entregada el 5 abr"),
    ("D-030", "2026-04-18", "Telefonia movil",        85.00, "out", "servicios",           "",          "completed", "Plan business Verizon"),
]
for i, row_data in enumerate(demo_txs, 2):
    for j, val in enumerate(row_data, 1):
        c = ws2.cell(row=i, column=j, value=val)
        c.font = Font(size=10, color=TEXT_PRIMARY)
        c.border = border
        if i % 2 == 0:
            c.fill = fill("F8FAFC")
        if j == 4:
            c.font = Font(size=10, bold=True, color=BLUE_TEXT)
            c.number_format = '"$"#,##0.00'
            c.alignment = Alignment(horizontal='right')
        elif j == 5:
            c.font = Font(size=10, bold=True, color=GREEN_TEXT if val == "in" else RED_TEXT)
            c.alignment = Alignment(horizontal='center')

widths = [10, 12, 24, 12, 11, 22, 18, 12, 40]
for i, w in enumerate(widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w


# =========================
# 4. CASHFLOW (con grafico mejorado)
# =========================
ws_cf = wb.create_sheet("Cashflow")
ws_cf.sheet_view.showGridLines = False

ws_cf['B2'] = "CASHFLOW DIARIO  ·  Mayo 2026"
ws_cf['B2'].font = Font(size=18, bold=True, color=GOLD)
ws_cf['B2'].fill = fill(DARK_BG)
ws_cf.merge_cells('B2:G2')
ws_cf['B2'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws_cf.row_dimensions[2].height = 38

ws_cf['B3'] = "Movimientos consolidados por fecha. Ingresos / Gastos / Flujo / Saldo acumulado."
ws_cf['B3'].font = Font(size=10, italic=True, color="64748B")
ws_cf.merge_cells('B3:G3')

cf_headers = ["Fecha", "Ingresos", "Gastos", "Flujo Neto", "Saldo Acumulado"]
for i, h in enumerate(cf_headers, 2):
    c = ws_cf.cell(row=5, column=i, value=h)
    c.font = Font(bold=True, color=GOLD, size=11)
    c.fill = fill(DARK_BG)
    c.alignment = Alignment(horizontal='center', vertical='center')
ws_cf.row_dimensions[5].height = 28

# Saldo inicial reference
ws_cf['I3'] = "Saldo inicial (Dashboard!E5):"
ws_cf['I3'].font = Font(size=9, italic=True, color="64748B")
ws_cf['J3'] = "=Dashboard!E5"
ws_cf['J3'].font = Font(size=10, bold=True, color=BLUE_TEXT)
ws_cf['J3'].number_format = '"$"#,##0.00'

dates = [
    "2026-04-18", "2026-04-19", "2026-04-20", "2026-04-21", "2026-04-22",
    "2026-04-23", "2026-04-24", "2026-04-25", "2026-04-26", "2026-04-27",
    "2026-04-28", "2026-04-29", "2026-04-30", "2026-05-01", "2026-05-02",
    "2026-05-03", "2026-05-04",
]
for idx, d in enumerate(dates):
    row = 6 + idx
    ws_cf.cell(row=row, column=2, value=d).font = Font(size=10, color=TEXT_PRIMARY)
    ws_cf.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    ws_cf.cell(row=row, column=3, value=f'=SUMIFS(Transacciones!D:D,Transacciones!E:E,"in",Transacciones!B:B,"{d}")')
    ws_cf.cell(row=row, column=3).font = Font(size=10, color=GREEN_TEXT)
    ws_cf.cell(row=row, column=3).number_format = '"$"#,##0.00'
    ws_cf.cell(row=row, column=4, value=f'=SUMIFS(Transacciones!D:D,Transacciones!E:E,"out",Transacciones!B:B,"{d}")')
    ws_cf.cell(row=row, column=4).font = Font(size=10, color=RED_TEXT)
    ws_cf.cell(row=row, column=4).number_format = '"$"#,##0.00'
    ws_cf.cell(row=row, column=5, value=f"=C{row}-D{row}")
    ws_cf.cell(row=row, column=5).font = Font(size=10, bold=True, color=DARK_BG)
    ws_cf.cell(row=row, column=5).number_format = '"$"#,##0.00'
    if idx == 0:
        ws_cf.cell(row=row, column=6, value=f"=$J$3+E{row}")
    else:
        ws_cf.cell(row=row, column=6, value=f"=F{row-1}+E{row}")
    ws_cf.cell(row=row, column=6).font = Font(size=10, bold=True, color=BLUE_TEXT)
    ws_cf.cell(row=row, column=6).number_format = '"$"#,##0.00'

totals_row = 6 + len(dates)
ws_cf.cell(row=totals_row, column=2, value="TOTAL PERIODO").font = Font(bold=True, color=DARK_BG, size=11)
ws_cf.cell(row=totals_row, column=2).fill = fill(GOLD_LIGHT)
ws_cf.cell(row=totals_row, column=2).alignment = Alignment(horizontal='center')
ws_cf.cell(row=totals_row, column=3, value=f"=SUM(C6:C{totals_row-1})").font = Font(bold=True, color=GREEN_TEXT, size=11)
ws_cf.cell(row=totals_row, column=3).number_format = '"$"#,##0.00'
ws_cf.cell(row=totals_row, column=4, value=f"=SUM(D6:D{totals_row-1})").font = Font(bold=True, color=RED_TEXT, size=11)
ws_cf.cell(row=totals_row, column=4).number_format = '"$"#,##0.00'
ws_cf.cell(row=totals_row, column=5, value=f"=C{totals_row}-D{totals_row}").font = Font(bold=True, color=DARK_BG, size=11)
ws_cf.cell(row=totals_row, column=5).number_format = '"$"#,##0.00'
ws_cf.cell(row=totals_row, column=6, value=f"=F{totals_row-1}").font = Font(bold=True, color=BLUE_TEXT, size=11)
ws_cf.cell(row=totals_row, column=6).number_format = '"$"#,##0.00'

# Chart fixes: no gridlines, wider, label skip
chart = LineChart()
chart.title = "Cashflow diario · Saldo acumulado"
chart.style = 12
chart.y_axis.title = "Monto USD"
chart.x_axis.title = "Fecha"
chart.height = 11
chart.width = 26

# Remove gridlines
chart.y_axis.majorGridlines = None
chart.x_axis.majorGridlines = None

# Reduce x-axis label density (skip every other date)
chart.x_axis.tickLblSkip = 2

# Number format for axes
chart.y_axis.number_format = '"$"#,##0'

data = Reference(ws_cf, min_col=5, min_row=5, max_col=6, max_row=5 + len(dates))
chart.add_data(data, titles_from_data=True)
cats = Reference(ws_cf, min_col=2, min_row=6, max_row=5 + len(dates))
chart.set_categories(cats)
ws_cf.add_chart(chart, "H6")

cf_widths = [3, 14, 14, 14, 14, 18, 4, 4, 30, 16]
for i, w in enumerate(cf_widths, 1):
    ws_cf.column_dimensions[get_column_letter(i)].width = w


# =========================
# 5. DEUDAS
# =========================
ws3 = wb.create_sheet("Deudas")
ws3.sheet_view.showGridLines = False

ws3['A1'] = "DEUDAS  ·  Por cobrar / Por pagar"
ws3['A1'].font = Font(size=14, bold=True, color=GOLD)
ws3['A1'].fill = fill(DARK_BG)
ws3.merge_cells('A1:N1')
ws3['A1'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws3.row_dimensions[1].height = 32

deudas_headers = [
    "Deudor / Acreedor", "Tipo", "Capital Original", "Fecha Inicio",
    "Tasa Mensual %", "Meses Transc.", "Interes Mensual Fijo", "Pagos a Capital",
    "Saldo Vivo", "Fecha Vcto.", "Dias p/Vcto.", "Forma de Pago",
    "Status", "Notas"
]
for i, h in enumerate(deudas_headers, 1):
    c = ws3.cell(row=3, column=i, value=h)
    c.font = Font(bold=True, color=DARK_BG, size=10)
    c.fill = fill(GOLD_LIGHT)
    c.alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[3].height = 26

deudas_data = [
    ("Luis Mendoza", "Por Pagar", 40000, "2025-12-15", 4, 4, 1600, 0,     40000, "", "", "Transferencia mensual", "Sin vcto.",
     "Prestamo dic 2025. Capital $40,000. Interes 4% mensual fijo $1,600. Solo se pagan intereses; capital intacto."),
    ("Pedro Ramos", "Por Pagar", 50000, "2025-12-15", 5, 4, 2500, 5000,   45000, "", "", "Transferencia mensual", "Activo",
     "Prestamo dic 2025. Capital $50K. Interes 5% mensual $2,500. Abono $5K capital el 29-abr-26. Saldo vivo $45K."),
]
for i, row_data in enumerate(deudas_data, 4):
    for j, val in enumerate(row_data, 1):
        c = ws3.cell(row=i, column=j, value=val)
        c.font = Font(size=10, color=TEXT_PRIMARY)
        if j == 3 or j == 7 or j == 8 or j == 9:
            c.font = Font(size=10, bold=True, color=BLUE_TEXT if j == 3 else LINK_COLOR if j in [7, 8] else TEXT_PRIMARY)
            c.number_format = '"$"#,##0.00'
            c.alignment = Alignment(horizontal='right')
        elif j == 5:
            c.number_format = '0.00\\%'
            c.alignment = Alignment(horizontal='center')

ws3.cell(row=10, column=1, value="TOTAL POR PAGAR").font = Font(bold=True, color=DARK_BG)
ws3.cell(row=10, column=1).fill = fill(GOLD_LIGHT)
ws3.cell(row=10, column=3, value=90000).font = Font(bold=True, size=11, color=BLUE_TEXT)
ws3.cell(row=10, column=3).number_format = '"$"#,##0.00'
ws3.cell(row=10, column=7, value=4100).font = Font(bold=True, size=11, color=LINK_COLOR)
ws3.cell(row=10, column=7).number_format = '"$"#,##0.00'
ws3.cell(row=10, column=8, value=5000).font = Font(bold=True, size=11, color=LINK_COLOR)
ws3.cell(row=10, column=8).number_format = '"$"#,##0.00'
ws3.cell(row=10, column=9, value=85000).font = Font(bold=True, size=11)
ws3.cell(row=10, column=9).number_format = '"$"#,##0.00'

# Credito rotativo
ws3.cell(row=14, column=1, value="CREDITO ROTATIVO PROVEEDOR").font = Font(size=12, bold=True, color=GOLD)
ws3.cell(row=14, column=1).fill = fill(DARK_BG)
ws3.merge_cells('A14:E14')

ws3.cell(row=15, column=1, value="Proveedor").font = Font(bold=True, color=DARK_BG, size=10)
ws3.cell(row=15, column=2, value="Cisterna Pendiente").font = Font(bold=True, color=DARK_BG, size=10)
ws3.cell(row=15, column=3, value="Monto").font = Font(bold=True, color=DARK_BG, size=10)
ws3.cell(row=15, column=4, value="Fecha Entrega").font = Font(bold=True, color=DARK_BG, size=10)
for col in [1, 2, 3, 4]:
    ws3.cell(row=15, column=col).fill = fill(GOLD_LIGHT)
    ws3.cell(row=15, column=col).alignment = Alignment(horizontal='center')

ws3.cell(row=16, column=1, value="Olivar Andaluz S.L.").font = Font(size=10, bold=True, color=BLUE_TEXT)
ws3.cell(row=16, column=2, value="Cisterna 18 (OL180)").font = Font(size=10, color=BLUE_TEXT)
ws3.cell(row=16, column=3, value=58420.50).font = Font(size=10, bold=True, color=BLUE_TEXT)
ws3.cell(row=16, column=3).number_format = '"$"#,##0.00'
ws3.cell(row=16, column=4, value="2026-05-22").font = Font(size=10, color=BLUE_TEXT)

ws3_widths = [22, 14, 14, 12, 14, 14, 18, 16, 16, 12, 12, 18, 14, 38]
for i, w in enumerate(ws3_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w


# =========================
# 6. KPIs
# =========================
ws4 = wb.create_sheet("KPIs")
ws4.sheet_view.showGridLines = False
ws4['B2'] = "KPIs (referencia tecnica)"
ws4['B2'].font = Font(size=16, bold=True, color=DARK_BG)

kpi_list = [
    ("Total Ingresos",        '=SUMIFS(Transacciones!D:D,Transacciones!E:E,"in")',  "money"),
    ("Total Gastos",          '=SUMIFS(Transacciones!D:D,Transacciones!E:E,"out")', "money"),
    ("Pagos a Proveedores",   '=SUMIFS(Transacciones!D:D,Transacciones!F:F,"pago_proveedor")', "money"),
    ("Pagos Deuda Interes",   '=SUMIFS(Transacciones!D:D,Transacciones!F:F,"pago_deuda_interes")', "money"),
    ("Pagos Deuda Capital",   '=SUMIFS(Transacciones!D:D,Transacciones!F:F,"pago_deuda_capital")', "money"),
    ("Cobros Clientes",       '=SUMIFS(Transacciones!D:D,Transacciones!F:F,"cobro_cliente")', "money"),
    ("Flujo Neto",            '=C5-C6',                                              "money"),
    ("# Cobros",              '=COUNTIFS(Transacciones!E:E,"in")',                  "count"),
    ("# Pagos",               '=COUNTIFS(Transacciones!E:E,"out")',                 "count"),
    ("Banco (calculado)",     '=Dashboard!C12',                                      "money"),
    ("Gandolas (calculado)",  '=Dashboard!E12',                                      "money"),
    ("Por Cobrar (calculado)",'=Dashboard!G12',                                      "money"),
    ("Total Trazado",         '=Dashboard!B8',                                       "money"),
]
for i, (label, formula, fmt) in enumerate(kpi_list, 5):
    ws4.cell(row=i, column=2, value=label).font = Font(size=11, color=TEXT_PRIMARY)
    c = ws4.cell(row=i, column=3, value=formula)
    c.font = Font(size=12, bold=True, color=BLUE_TEXT)
    c.number_format = '#,##0' if fmt == "count" else '"$"#,##0.00'

ws4.column_dimensions['B'].width = 32
ws4.column_dimensions['C'].width = 22


output_path = os.path.join("posts", "demo-sheet-for-loom.xlsx")
wb.save(output_path)
print("Saved:", output_path)
print("Size:", os.path.getsize(output_path), "bytes")
