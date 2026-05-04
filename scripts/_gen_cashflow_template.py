from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DARK_BG = "0F172A"
GOLD = "D4B88C"
INPUT_COLOR = "2563EB"
TEXT_PRIMARY = "1A1A1A"
TEXT_MUTED = "94A3B8"

wb = Workbook()
thin = Side(border_style="thin", color="E2E8F0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# === PORTADA ===
ws = wb.active
ws.title = "Portada"
ws.sheet_view.showGridLines = False

ws['B2'] = "CASHFLOW TEMPLATE"
ws['B2'].font = Font(size=22, bold=True, color=DARK_BG)
ws['B3'] = "Para importadoras y PYMEs que reciben Zelle / transferencias"
ws['B3'].font = Font(size=13, italic=True, color="64748B")

ws['B5'] = "Powered by ZENIA  -  zeniapartners.com"
ws['B5'].font = Font(size=11, color=INPUT_COLOR)

ws['B7'] = "Como usar este template"
ws['B7'].font = Font(size=14, bold=True, color=DARK_BG)

instructions = [
    "1. Hacer una copia a tu Google Drive (Archivo, Hacer una copia)",
    "2. Ir a la pestana Transacciones y registrar cada Zelle/transferencia",
    "3. Categorizar cada movimiento (cobro_cliente, gasto_logistica, etc)",
    "4. Volver a esta portada para ver tus 3 buckets de efectivo automaticos",
    "5. Editar el saldo inicial banco en E18 con tu saldo real al inicio",
]
for i, txt in enumerate(instructions):
    ws.cell(row=8 + i, column=2, value=txt).font = Font(size=11, color=TEXT_PRIMARY)

ws['B15'] = "TRAZABILIDAD DEL EFECTIVO (calculada automaticamente)"
ws['B15'].font = Font(size=12, bold=True, color=GOLD)
ws['B15'].fill = PatternFill("solid", fgColor=DARK_BG)
ws.merge_cells('B15:H15')
ws['B15'].alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[15].height = 30

cards = [
    ('B17', 'EFECTIVO EN BANCO', 'B18',
     '=E18+SUMIFS(Transacciones!D:D,Transacciones!E:E,"in")-SUMIFS(Transacciones!D:D,Transacciones!E:E,"out")-(SUMIFS(Transacciones!D:D,Transacciones!F:F,"entrega_gandola")-SUMIFS(Transacciones!D:D,Transacciones!F:F,"retorno_gandola"))-(SUMIFS(Transacciones!D:D,Transacciones!F:F,"venta_credito")-SUMIFS(Transacciones!D:D,Transacciones!F:F,"cobro_credito"))',
     'B19', 'Saldo liquido disponible', "DBEAFE"),
    ('D17', 'DINERO EN GANDOLAS', 'D18',
     '=SUMIFS(Transacciones!D:D,Transacciones!F:F,"entrega_gandola")-SUMIFS(Transacciones!D:D,Transacciones!F:F,"retorno_gandola")',
     'D19', 'Capital en transito', "FEF3C7"),
    ('F17', 'CUENTAS POR COBRAR', 'F18',
     '=SUMIFS(Transacciones!D:D,Transacciones!F:F,"venta_credito")-SUMIFS(Transacciones!D:D,Transacciones!F:F,"cobro_credito")',
     'F19', 'Pendiente de cobrar', "FED7AA"),
    ('H17', 'TOTAL TRAZADO', 'H18', '=B18+D18+F18',
     'H19', 'Suma de los 3', "DCFCE7"),
]
for label_cell, label, val_cell, formula, sub_cell, sub, bg in cards:
    ws[label_cell] = label
    ws[label_cell].font = Font(size=10, bold=True, color=DARK_BG)
    ws[label_cell].fill = PatternFill("solid", fgColor=bg)
    ws[label_cell].alignment = Alignment(horizontal='center', vertical='center')

    ws[val_cell] = formula
    ws[val_cell].font = Font(size=18, bold=True, color=INPUT_COLOR)
    ws[val_cell].fill = PatternFill("solid", fgColor=bg)
    ws[val_cell].alignment = Alignment(horizontal='center', vertical='center')
    ws[val_cell].number_format = '"$"#,##0.00'

    ws[sub_cell] = sub
    ws[sub_cell].font = Font(size=9, italic=True, color="64748B")
    ws[sub_cell].fill = PatternFill("solid", fgColor=bg)
    ws[sub_cell].alignment = Alignment(horizontal='center')

ws.row_dimensions[17].height = 22
ws.row_dimensions[18].height = 50
ws.row_dimensions[19].height = 18

ws['B21'] = "SALDO INICIAL BANCO (editable):"
ws['B21'].font = Font(size=11, italic=True, color="64748B")
ws.merge_cells('B21:D21')
ws['B21'].alignment = Alignment(horizontal='right')
ws['E18'] = 0
ws['E18'].font = Font(size=11, bold=True, color=INPUT_COLOR)
ws['E18'].number_format = '"$"#,##0.00'
ws['E18'].fill = PatternFill("solid", fgColor="DBEAFE")

# Column widths
for col, w in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I'],
                  [3, 24, 24, 24, 24, 24, 24, 24, 3]):
    ws.column_dimensions[col].width = w

# Footer Portada
ws['B30'] = "Quieres que esto trabaje SOLO?"
ws['B30'].font = Font(size=14, bold=True, color=DARK_BG)
ws['B31'] = "Manda capturas a Gaia (asistente IA por WhatsApp), ella categoriza y registra."
ws['B31'].font = Font(size=11, color=TEXT_PRIMARY)
ws['B32'] = "Caso: zeniapartners.com/cases/whatsapp-bookkeeping-importer.html"
ws['B32'].font = Font(size=11, color=INPUT_COLOR, underline='single')
ws['B33'] = "Agenda 30 min: calendly.com/zeladauriartef/30min"
ws['B33'].font = Font(size=11, color=INPUT_COLOR, underline='single')

# === TRANSACCIONES ===
ws2 = wb.create_sheet("Transacciones")
ws2.sheet_view.showGridLines = False

headers = ["ID", "Fecha", "Nombre", "Monto", "Direccion", "Categoria", "Conf#", "Estado", "Notas"]
for i, h in enumerate(headers, 1):
    c = ws2.cell(row=1, column=i, value=h)
    c.font = Font(bold=True, color=GOLD, size=11)
    c.fill = PatternFill("solid", fgColor=DARK_BG)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border
ws2.row_dimensions[1].height = 32

samples = [
    ("EJ-001", "2026-05-01", "Carlos Perez", 500.00, "in", "cobro_cliente", "ABC123XYZ", "completed", "Pago anticipo"),
    ("EJ-002", "2026-05-02", "Gandola Beleva", 5000.00, "out", "entrega_gandola", "", "completed", "Capital viaje Caracas"),
    ("EJ-003", "2026-05-03", "Juan Rodriguez", 800.00, "out", "venta_credito", "", "completed", "Mercancia paga el 15"),
]
for i, row_data in enumerate(samples, 2):
    for j, val in enumerate(row_data, 1):
        c = ws2.cell(row=i, column=j, value=val)
        c.font = Font(size=10, italic=True, color=TEXT_MUTED)
        c.border = border

widths = [12, 12, 22, 12, 11, 22, 14, 12, 28]
for i, w in enumerate(widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.cell(row=6, column=2,
         value="Filas de ejemplo. Borralas y empieza a registrar las tuyas.").font = Font(
    size=10, bold=True, italic=True, color="DC2626")

# === GUIA CATEGORIAS ===
ws3 = wb.create_sheet("Guia Categorias")
ws3.sheet_view.showGridLines = False

ws3['B2'] = "GUIA DE CATEGORIAS"
ws3['B2'].font = Font(size=18, bold=True, color=DARK_BG)
ws3['B3'] = "16 categorias contables para importadoras y PYMEs que reciben Zelle"
ws3['B3'].font = Font(size=11, italic=True, color="64748B")

cats = [
    ("ENTRADAS (in)", "", "", DARK_BG, GOLD),
    ("cobro_cliente", "Cliente paga por Zelle/transferencia",
     "Mejora Banco. Categoria mas comun.", "DBEAFE", "1E40AF"),
    ("retorno_gandola", "Capital o efectivo que vuelve de una gandola",
     "Reduce Gandolas, aumenta Banco.", "DCFCE7", "166534"),
    ("cobro_credito", "Cliente paga lo que debia",
     "Reduce Cuentas por Cobrar, aumenta Banco.", "DCFCE7", "166534"),
    ("cobro_deuda_capital", "Te devuelven capital de un prestamo Por Cobrar",
     "Reduce Saldo Vivo del prestamo.", "EDE9FE", "5B21B6"),
    ("cobro_deuda_interes", "Te pagan intereses de un prestamo Por Cobrar",
     "Solo registro, no toca capital.", "EDE9FE", "5B21B6"),
    ("SALIDAS (out)", "", "", DARK_BG, GOLD),
    ("gasto_logistica", "Transporte, fletes, distribucion", "Reduce Banco.", "FEE2E2", "991B1B"),
    ("planilla", "Pagos de salarios fijos", "Reduce Banco. Mensual recurrente.", "FEE2E2", "991B1B"),
    ("impuesto_gandola", "Aduana, peajes, permisos", "Reduce Banco.", "FEE2E2", "991B1B"),
    ("viaticos", "Gastos de viaje, alimentacion en ruta", "Reduce Banco.", "FEE2E2", "991B1B"),
    ("intereses_prestamo", "Intereses mensuales de un prestamo",
     "Reduce Banco. Mensual fijo.", "FEF3C7", "92400E"),
    ("pago_deuda_capital", "Abono al capital de un prestamo Por Pagar",
     "Reduce Banco y reduce Saldo Vivo.", "FEF3C7", "92400E"),
    ("pago_deuda_interes", "Pago intereses mensual a un prestamista",
     "Reduce Banco, no toca capital.", "FEF3C7", "92400E"),
    ("entrega_gandola", "Capital entregado a una gandola activa",
     "Reduce Banco, aumenta Gandolas.", "FED7AA", "9A3412"),
    ("venta_credito", "Mercancia entregada sin cobrar todavia",
     "No toca Banco. Aumenta Cuentas por Cobrar.", "FED7AA", "9A3412"),
    ("fee_transaccion", "Comisiones bancarias o de plataforma",
     "Reduce Banco.", "FEE2E2", "991B1B"),
    ("sin_clasificar", "Cuando no encaja en ninguna otra",
     "Catch-all temporal, reclasificar luego.", "F1F5F9", "475569"),
]

row = 5
for cat in cats:
    name, desc, effect, bg, fg = cat
    if not desc and not effect:
        ws3.cell(row=row, column=2, value=name).font = Font(size=12, bold=True, color=fg)
        ws3.cell(row=row, column=2).fill = PatternFill("solid", fgColor=bg)
        ws3.merge_cells(f'B{row}:E{row}')
        ws3.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center')
        ws3.row_dimensions[row].height = 24
    else:
        ws3.cell(row=row, column=2, value=name).font = Font(size=11, bold=True, color=fg)
        ws3.cell(row=row, column=2).fill = PatternFill("solid", fgColor=bg)
        ws3.cell(row=row, column=3, value=desc).font = Font(size=10, color=TEXT_PRIMARY)
        ws3.cell(row=row, column=3).fill = PatternFill("solid", fgColor=bg)
        ws3.cell(row=row, column=4, value=effect).font = Font(size=10, italic=True, color="64748B")
        ws3.cell(row=row, column=4).fill = PatternFill("solid", fgColor=bg)
        ws3.row_dimensions[row].height = 24
    row += 1

ws3.column_dimensions['B'].width = 28
ws3.column_dimensions['C'].width = 50
ws3.column_dimensions['D'].width = 50

# === KPIs ===
ws4 = wb.create_sheet("KPIs")
ws4.sheet_view.showGridLines = False
ws4['B2'] = "KPIs DEL PERIODO"
ws4['B2'].font = Font(size=16, bold=True, color=DARK_BG)

kpis = [
    ("Total Ingresos", '=SUMIFS(Transacciones!D:D,Transacciones!E:E,"in")', "money"),
    ("Total Gastos", '=SUMIFS(Transacciones!D:D,Transacciones!E:E,"out")', "money"),
    ("Flujo Neto", '=C5-C6', "money"),
    ("# Transferencias recibidas", '=COUNTIFS(Transacciones!E:E,"in")', "count"),
    ("# Transferencias enviadas", '=COUNTIFS(Transacciones!E:E,"out")', "count"),
    ("Ticket promedio recibido", '=IFERROR(C5/C8,0)', "money"),
    ("Capital en gandolas", '=Portada!D18', "money"),
    ("Cuentas por cobrar", '=Portada!F18', "money"),
]
for i, (label, formula, fmt) in enumerate(kpis, 5):
    ws4.cell(row=i, column=2, value=label).font = Font(size=11, color=TEXT_PRIMARY)
    c = ws4.cell(row=i, column=3, value=formula)
    c.font = Font(size=12, bold=True, color=INPUT_COLOR)
    c.number_format = '#,##0' if fmt == "count" else '"$"#,##0.00'

ws4.column_dimensions['B'].width = 32
ws4.column_dimensions['C'].width = 22

import os
output_path = os.path.join("lead-magnets", "cashflow-importer-template.xlsx")
wb.save(output_path)
print("Saved:", output_path)
print("Size:", os.path.getsize(output_path), "bytes")
